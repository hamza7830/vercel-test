import json
import os
import re
import sys
from http.server import BaseHTTPRequestHandler
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Add project root to Python path (need to go up 4 levels from vercel/api/generate-copy/seo/)
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))
sys.path.insert(0, project_root)

# Now you can import from _app / _helper (these must exist in your repo!)
try:
    from _app.local.environment import is_running_locally, load_env
    from _helper.one_password import get_json_note_sync
except ImportError as ie:
    # Log and skip if not available (for deployment/debug)
    sys.stderr.write(f"Could not import local helpers: {ie}\n")

def check_dependencies():
    """
    Check if required dependencies are available.
    Logs the import and version info for each dependency.
    """
    debug("=== Checking Dependencies ===")

    deps = [
        ("openpyxl", "openpyxl")
    ]

    for dep_name, import_name in deps:
        try:
            mod = __import__(import_name)
            version = getattr(mod, "__version__", "unknown")
            debug(f"✅ {dep_name} is available (version: {version})")
        except ImportError as e:
            debug(f"❌ {dep_name} import failed: {e}")
        except Exception as e:
            debug(f"⚠️ {dep_name} import error: {e}")

    debug("=== End Dependencies Check ===")

def debug(msg):
    sys.stderr.write(msg + "\n")
    sys.stderr.flush()

def debug_environment():
    """Debug all environment variables"""
    debug("=== Environment Variables ===")
    for key, value in os.environ.items():
        debug(f"{key}: {value}")
    debug("=== End Environment Variables ===")

def explore_filesystem():
    """Explore the filesystem to see what's available"""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    debug(f"Current handler location: {current_dir}")
    for i in range(5):
        test_path = os.path.join(current_dir, *(['..'] * i))
        abs_path = os.path.abspath(test_path)
        debug(f"Level {i} ({abs_path}): {os.listdir(abs_path) if os.path.exists(abs_path) else 'Not found'}")
        app_path = os.path.join(abs_path, '_app')
        if os.path.exists(app_path):
            debug(f"Found _app at level {i}: {os.listdir(app_path)}")
            local_path = os.path.join(app_path, 'local')
            if os.path.exists(local_path):
                debug(f"Found _app/local: {os.listdir(local_path)}")

def parse_seo_output(llm_text: str):
    rows = []
    current_url = None
    current_brand = None
    title = None
    meta = None
    title_count = None
    meta_count = None
    lines = llm_text.strip().splitlines()
    for line in lines:
        line = line.strip()
        m = re.match(r"^Line \d+ \(URL: ([^)]+)\):", line)
        if m:
            current_url = m.group(1).strip()
            continue
        m = re.match(r"^For input:.*Brand\s*{([^}-]+)", line)
        if m:
            current_brand = m.group(1).strip()
            continue
        m = re.match(r"^Title \d+:\s*(.+?)(?:\s*\((\d+)\))?$", line)
        if m:
            title = m.group(1).strip()
            title_count = int(m.group(2)) if m.group(2) else len(title)
            continue
        m = re.match(r"^Meta Description \d+:\s*(.+?)(?:\s*\((\d+)\))?$", line)
        if m:
            meta = m.group(1).strip()
            meta_count = int(m.group(2)) if m.group(2) else len(meta)
            if current_url and current_brand and title and meta:
                rows.append((current_brand, current_url, title, title_count, meta, meta_count))
                title = None
                meta = None
                title_count = None
                meta_count = None
            continue
    return rows

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        debug("SEO XLSX handler started - GET")

        # Demo: filesystem and environment diagnostics (uncomment if needed)
        # explore_filesystem()
        # debug_environment()

        # Only load env if running locally
        try:
            if 'is_running_locally' in globals() and is_running_locally():
                load_env()
            else:
                debug("Running in cloud environment - skipping local env loading")
        except Exception as e:
            debug(f"Env load error: {e}")

        check_dependencies()

        try:
            if 'get_json_note_sync' in globals():
                get_json_note_sync("eftj3nyjzwx6xf4mpjdep4jmpu", "bnl2n2hc6kldzgnusvrae3lbcy")
        except Exception as e:
            debug(f"get_json_note_sync error: {e}")

        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Type', 'application/json')
        self.end_headers()

        response_data = {
            "message": "SEO - XLSX - Export Done!",
            "method": "GET",
            "environment": {
                "EMS_ENVIRONMENT": os.environ.get('EMS_ENVIRONMENT', 'not set'),
                "VERCEL_ENV": os.environ.get('VERCEL_ENV', 'not set')
            },
            "timestamp": str(os.environ.get('VERCEL_REQUEST_TIME', 'not available'))
        }

        debug(f"Returning response: {response_data}")
        self.wfile.write(json.dumps(response_data, indent=2).encode())

    def do_POST(self):
        debug("SEO XLSX handler started - POST")
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        try:
            data = json.loads(post_data.decode())
        except Exception:
            self.send_response(400)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": "Invalid JSON"}
            self.wfile.write(json.dumps(error).encode())
            return

        llm_output = data.get("llm_output")
        if not llm_output:
            self.send_response(400)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": "No llm_output provided"}
            self.wfile.write(json.dumps(error).encode())
            return

        # Path relative to this file's location
        script_dir = os.path.dirname(__file__)
        template_path = os.path.join(script_dir, "../resources/templates/template_seo.xlsx")
        image_path = os.path.join(script_dir, "../resources/images/8ms.png")

        try:
            wb = load_workbook(template_path)
            ws = wb.active

            try:
                img = Image(image_path)
                img.width = 500
                img.height = 77.45
                ws.add_image(img, 'B2')
                debug("Image added to workbook")
            except Exception as img_e:
                debug(f"Image error: {img_e}")

            rows = parse_seo_output(llm_output)
            if len(rows) == 0:
                self.send_response(400)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                error = {"error": "No valid SEO metadata parsed."}
                self.wfile.write(json.dumps(error).encode())
                return

            start_row = 10
            for idx, (brand, url, title, title_count, meta, meta_count) in enumerate(rows):
                row = start_row + idx
                ws.cell(row=row, column=2, value=brand)
                ws.cell(row=row, column=3, value=url)
                ws.cell(row=row, column=4, value=title)
                ws.cell(row=row, column=5, value=title_count)
                ws.cell(row=row, column=6, value=meta)
                ws.cell(row=row, column=7, value=meta_count)
            debug("Excel cells filled")

            output = BytesIO()
            wb.save(output)
            file_bytes = output.getvalue()
            debug(f"Workbook saved to BytesIO, length: {len(file_bytes)} bytes")

            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename=SEO_Metadata.xlsx')
            self.end_headers()
            self.wfile.write(file_bytes)
        except Exception as e:
            self.send_response(500)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": str(e)}
            self.wfile.write(json.dumps(error).encode())

    def do_OPTIONS(self):
        debug("Handling OPTIONS preflight")
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
