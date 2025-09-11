import json
import os
import re
import sys
from http.server import BaseHTTPRequestHandler
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Add project root to Python path (need to go up 4 levels from vercel/api/generate-copy/facebook/)
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))
sys.path.insert(0, project_root)

# Try importing helpers if available
try:
    from _app.local.environment import is_running_locally, load_env
    from _helper.one_password import get_json_note_sync
except ImportError as ie:
    sys.stderr.write(f"Could not import local helpers: {ie}\n")

def debug(msg):
    sys.stderr.write(msg + "\n")
    sys.stderr.flush()

def check_dependencies():
    debug("=== Checking Dependencies ===")
    deps = [
        ("openpyxl", "openpyxl"),
    ]
    for dep_name, import_name in deps:
        try:
            mod = __import__(import_name)
            version = getattr(mod, "__version__", "unknown")
            debug(f"✅ {dep_name} is available (version: {version})")
        except ImportError as e:
            debug(f"❌ {dep_name} import failed: {e}")
        except Exception as e:
            debug(f"⚠️ {dep_name} import error: {e}")
    debug("=== End Dependencies Check ===")

def debug_environment():
    debug("=== Environment Variables ===")
    for key, value in os.environ.items():
        debug(f"{key}: {value}")
    debug("=== End Environment Variables ===")

def explore_filesystem():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    debug(f"Current handler location: {current_dir}")
    for i in range(5):
        test_path = os.path.join(current_dir, *(['..'] * i))
        abs_path = os.path.abspath(test_path)
        debug(f"Level {i} ({abs_path}): {os.listdir(abs_path) if os.path.exists(abs_path) else 'Not found'}")
        app_path = os.path.join(abs_path, '_app')
        if os.path.exists(app_path):
            debug(f"Found _app at level {i}: {os.listdir(app_path)}")
            local_path = os.path.join(app_path, 'local')
            if os.path.exists(local_path):
                debug(f"Found _app/local: {os.listdir(local_path)}")

def extract_text_and_count(text):
    t = text.strip()
    return t, len(t)

def parse_llm_output(llm_text: str):
    data_rows = []
    current_channel = None
    primary_text = None
    headline = None

    lines = llm_text.strip().splitlines()
    for line in lines:
        line = line.strip()
        if not line or re.match(r"^[-]{2,}$", line):
            continue

        channel_match = re.match(
            r"^(?:\s*[\*#]+\s*)?(?:\d+\.\s*)?(?:\*+)?\s*(Image Facebook Feed|Facebook Stories|Facebook Reels|Facebook Video Feed)\s*(?:\*+)?$",
            line, re.I
        )
        if channel_match:
            current_channel = channel_match.group(1).strip()
            continue

        primary_match = re.match(r"^Primary text:\s*(.+)", line, re.I)
        if primary_match:
            primary_text = primary_match.group(1).strip()
            continue

        headline_match = re.match(r"^Headline:\s*(.+)", line, re.I)
        if headline_match:
            headline = headline_match.group(1).strip()
            data_rows.append({
                "channel": current_channel or "",
                "primary_text": primary_text or "",
                "headline": headline or ""
            })
            primary_text = None
            headline = None

    return data_rows

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        debug("Facebook Ads XLSX handler started - GET")
        # Uncomment if you want diagnostics
        # explore_filesystem()
        # debug_environment()
        try:
            if 'is_running_locally' in globals() and is_running_locally():
                load_env()
            else:
                debug("Running in cloud environment - skipping local env loading")
        except Exception as e:
            debug(f"Env load error: {e}")

        check_dependencies()

        try:
            if 'get_json_note_sync' in globals():
                get_json_note_sync("eftj3nyjzwx6xf4mpjdep4jmpu", "bnl2n2hc6kldzgnusvrae3lbcy")
        except Exception as e:
            debug(f"get_json_note_sync error: {e}")

        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Type', 'application/json')
        self.end_headers()

        response_data = {
            "message": "Facebook Ads - XLSX - Export Done!",
            "method": "GET",
            "environment": {
                "EMS_ENVIRONMENT": os.environ.get('EMS_ENVIRONMENT', 'not set'),
                "VERCEL_ENV": os.environ.get('VERCEL_ENV', 'not set')
            },
            "timestamp": str(os.environ.get('VERCEL_REQUEST_TIME', 'not available'))
        }
        debug(f"Returning response: {response_data}")
        self.wfile.write(json.dumps(response_data, indent=2).encode())

    def do_POST(self):
        debug("Facebook Ads XLSX handler started - POST")
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        try:
            data = json.loads(post_data.decode())
        except Exception:
            self.send_response(400)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": "Invalid JSON"}
            self.wfile.write(json.dumps(error).encode())
            return

        llm_output = data.get("llm_output")
        if not llm_output:
            self.send_response(400)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": "No llm_output provided"}
            self.wfile.write(json.dumps(error).encode())
            return

        script_dir = os.path.dirname(__file__)
        template_path = os.path.join(script_dir, "../resources/templates/template_ad_copy.xlsx")
        image_path = os.path.join(script_dir, "../resources/image/8ms.png")

        try:
            wb = load_workbook(template_path)

            # Check for 'Ad Copy' sheet and select it
            if "Ad Copy" not in wb.sheetnames:
                ws = wb.create_sheet("Ad Copy")
            else:
                ws = wb["Ad Copy"]

            try:
                img = Image(image_path)
                img.width = 500
                img.height = 77.45
                ws.add_image(img, 'B2')
                debug("Image added to workbook")
            except Exception as img_e:
                debug(f"Error adding image: {img_e}. Skipping image addition.")

            rows = parse_llm_output(llm_output)
            if len(rows) == 0:
                self.send_response(400)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                error = {"error": "No valid ad copy parsed."}
                self.wfile.write(json.dumps(error).encode())
                return

            start_row = 10
            for idx, row in enumerate(rows):
                excel_row = start_row + idx
                channel = row.get("channel", "")
                primary = row.get("primary_text", "")
                headline = row.get("headline", "")

                headline_text, headline_count = extract_text_and_count(headline)
                primary_text, primary_count = extract_text_and_count(primary)

                ws.cell(row=excel_row, column=2, value=channel)       # Column B: Channel
                ws.cell(row=excel_row, column=6, value=headline_text)    # Column F: Headline
                ws.cell(row=excel_row, column=7, value=headline_count) # Column G: Headline char count
                ws.cell(row=excel_row, column=8, value=primary_text)    # Column H: Primary Text
                ws.cell(row=excel_row, column=9, value=primary_count)   # Column I: Primary Text char count

            output = BytesIO()
            wb.save(output)
            file_bytes = output.getvalue()
            debug(f"Workbook saved to BytesIO, length: {len(file_bytes)} bytes")

            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename=FacebookCopyFilled.xlsx')
            self.end_headers()
            self.wfile.write(file_bytes)
        except Exception as e:
            self.send_response(500)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": str(e)}
            self.wfile.write(json.dumps(error).encode())

    def do_OPTIONS(self):
        debug("Handling OPTIONS preflight")
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()