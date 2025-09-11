import json
import os
import re
import sys
from http.server import BaseHTTPRequestHandler
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Add project root to Python path (need to go up 4 levels from vercel/api/generate-copy/google/)
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))
sys.path.insert(0, project_root)

# Try importing helpers if available
try:
    from _app.local.environment import is_running_locally, load_env
    from _helper.one_password import get_json_note_sync
except ImportError as ie:
    sys.stderr.write(f"Could not import local helpers: {ie}\n")

def debug(msg):
    sys.stderr.write(msg + "\n")
    sys.stderr.flush()

def check_dependencies():
    debug("=== Checking Dependencies ===")
    deps = [
        ("openpyxl", "openpyxl"),
    ]
    for dep_name, import_name in deps:
        try:
            mod = __import__(import_name)
            version = getattr(mod, "__version__", "unknown")
            debug(f"✅ {dep_name} is available (version: {version})")
        except ImportError as e:
            debug(f"❌ {dep_name} import failed: {e}")
        except Exception as e:
            debug(f"⚠️ {dep_name} import error: {e}")
    debug("=== End Dependencies Check ===")

def debug_environment():
    debug("=== Environment Variables ===")
    for key, value in os.environ.items():
        debug(f"{key}: {value}")
    debug("=== End Environment Variables ===")

def explore_filesystem():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    debug(f"Current handler location: {current_dir}")
    for i in range(5):
        test_path = os.path.join(current_dir, *(['..'] * i))
        abs_path = os.path.abspath(test_path)
        debug(f"Level {i} ({abs_path}): {os.listdir(abs_path) if os.path.exists(abs_path) else 'Not found'}")
        app_path = os.path.join(abs_path, '_app')
        if os.path.exists(app_path):
            debug(f"Found _app at level {i}: {os.listdir(app_path)}")
            local_path = os.path.join(app_path, 'local')
            if os.path.exists(local_path):
                debug(f"Found _app/local: {os.listdir(local_path)}")

def extract_text_and_count(text):
    t = text.strip()
    return t, len(t)

def parse_google_ads_output(llm_text: str):
    FIELD_BASE = [
        "Headline (1)", "Headline (2)",
        "Description (1)", "Description (2)",
        "Path (1)", "Path (2)"
    ]
    data_rows = []
    sitelinks_data = []
    ad_block = {}
    ad_index = 0
    lines = llm_text.strip().splitlines()
    for line in lines:
        line = line.strip()
        if not line or re.match(r"^[-]{2,}$", line):
            continue
        matched = False
        for field in FIELD_BASE:
            m = re.match(rf"^{re.escape(field)}:\s*(.+?)(?:\s*\(\d+\))?$", line, re.I)
            if m:
                ad_block[field] = m.group(1).strip()
                matched = True
                break
        if not matched:
            for i in range(1, 21):
                text_key = f"SiteLink ({i})"
                desc_key = f"SiteLink Description ({i})"
                url_key = f"SiteLink URL ({i})"
                if line.startswith(text_key):
                    ad_block.setdefault("sitelinks", {}).setdefault(i, {})["text"] = line.split(":", 1)[1].strip()
                    matched = True
                elif line.startswith(desc_key):
                    ad_block.setdefault("sitelinks", {}).setdefault(i, {})["description"] = line.split(":", 1)[1].strip()
                    matched = True
                elif line.startswith(url_key):
                    ad_block.setdefault("sitelinks", {}).setdefault(i, {})["url"] = line.split(":", 1)[1].strip()
                    matched = True
        if all(field in ad_block for field in FIELD_BASE):
            row = {f: ad_block.get(f, "") for f in FIELD_BASE}
            data_rows.append(row)
            sitelinks = ad_block.get("sitelinks", {})
            for i, values in sitelinks.items():
                sitelinks_data.append({
                    "ad_index": ad_index,
                    "index": i,
                    "text": values.get("text", ""),
                    "description": values.get("description", ""),
                    "url": values.get("url", ""),
                })
            ad_index += 1
            ad_block = {}
    return data_rows, sitelinks_data

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        debug("Google Ads XLSX handler started - GET")
        # Uncomment if you want diagnostics
        # explore_filesystem()
        # debug_environment()
        try:
            if 'is_running_locally' in globals() and is_running_locally():
                load_env()
            else:
                debug("Running in cloud environment - skipping local env loading")
        except Exception as e:
            debug(f"Env load error: {e}")

        check_dependencies()

        try:
            if 'get_json_note_sync' in globals():
                get_json_note_sync("eftj3nyjzwx6xf4mpjdep4jmpu", "bnl2n2hc6kldzgnusvrae3lbcy")
        except Exception as e:
            debug(f"get_json_note_sync error: {e}")

        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Type', 'application/json')
        self.end_headers()

        response_data = {
            "message": "Google Ads - XLSX - Export Done!",
            "method": "GET",
            "environment": {
                "EMS_ENVIRONMENT": os.environ.get('EMS_ENVIRONMENT', 'not set'),
                "VERCEL_ENV": os.environ.get('VERCEL_ENV', 'not set')
            },
            "timestamp": str(os.environ.get('VERCEL_REQUEST_TIME', 'not available'))
        }
        debug(f"Returning response: {response_data}")
        self.wfile.write(json.dumps(response_data, indent=2).encode())

    def do_POST(self):
        debug("Google Ads XLSX handler started - POST")
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        try:
            data = json.loads(post_data.decode())
        except Exception:
            self.send_response(400)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": "Invalid JSON"}
            self.wfile.write(json.dumps(error).encode())
            return

        llm_output = data.get("llm_output")
        if not llm_output:
            self.send_response(400)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": "No llm_output provided"}
            self.wfile.write(json.dumps(error).encode())
            return

        script_dir = os.path.dirname(__file__)
        template_path = os.path.join(script_dir, "../resources/templates/template_ad_copy.xlsx")
        image_path = os.path.join(script_dir, "../resources/image/8ms.png")

        try:
            wb = load_workbook(template_path)

            # Main worksheet
            if "Ad Copy" not in wb.sheetnames:
                ws_main = wb.create_sheet("Ad Copy")
            else:
                ws_main = wb["Ad Copy"]

            try:
                img_main = Image(image_path)
                img_main.width = 500
                img_main.height = 77.45
                ws_main.add_image(img_main, 'B2')
                debug("Image added to 'Ad Copy'.")
            except Exception as img_e:
                debug(f"Image error: {img_e}")

            # Sitelinks worksheet
            if "Sitelinks" in wb.sheetnames:
                ws_sitelinks = wb["Sitelinks"]
            else:
                ws_sitelinks = wb.create_sheet("Sitelinks")
            try:
                img_sitelinks = Image(image_path)
                img_sitelinks.width = 500
                img_sitelinks.height = 77.45
                ws_sitelinks.add_image(img_sitelinks, 'B2')
                debug("Image added to 'Sitelinks'.")
            except Exception as img_e:
                debug(f"Image error: {img_e}")

            rows, sitelinks = parse_google_ads_output(llm_output)
            if len(rows) == 0:
                self.send_response(400)
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                error = {"error": "No valid ad copy parsed."}
                self.wfile.write(json.dumps(error).encode())
                return

            start_row = 10
            for idx, row in enumerate(rows):
                row1 = start_row + idx * 2
                row2 = row1 + 1

                h1_text, h1_count = extract_text_and_count(row.get("Headline (1)", ""))
                h2_text, h2_count = extract_text_and_count(row.get("Headline (2)", ""))
                d1_text, d1_count = extract_text_and_count(row.get("Description (1)", ""))
                d2_text, d2_count = extract_text_and_count(row.get("Description (2)", ""))
                p1_text, _ = extract_text_and_count(row.get("Path (1)", ""))
                p2_text, _ = extract_text_and_count(row.get("Path (2)", ""))

                ws_main.cell(row=row1, column=2, value="Google Ads")
                ws_main.cell(row=row2, column=2, value="Google Ads")

                ws_main.cell(row=row1, column=6, value=h1_text)
                ws_main.cell(row=row1, column=7, value=h1_count)
                ws_main.cell(row=row1, column=5, value=p1_text)
                ws_main.cell(row=row1, column=8, value=d1_text)
                ws_main.cell(row=row1, column=9, value=d1_count)

                ws_main.cell(row=row2, column=6, value=h2_text)
                ws_main.cell(row=row2, column=7, value=h2_count)
                ws_main.cell(row=row2, column=5, value=p2_text)
                ws_main.cell(row=row2, column=8, value=d2_text)
                ws_main.cell(row=row2, column=9, value=d2_count)

            sitelink_start_row = 10
            for idx, entry in enumerate(sitelinks):
                row = sitelink_start_row + idx
                text, text_count = extract_text_and_count(entry["text"])
                desc, desc_count = extract_text_and_count(entry["description"])
                ws_sitelinks.cell(row=row, column=2, value=text)
                ws_sitelinks.cell(row=row, column=3, value=text_count)
                ws_sitelinks.cell(row=row, column=4, value=desc)
                ws_sitelinks.cell(row=row, column=5, value=desc_count)
                ws_sitelinks.cell(row=row, column=8, value=entry["url"])

            output = BytesIO()
            wb.save(output)
            file_bytes = output.getvalue()
            debug(f"Workbook saved to BytesIO, length: {len(file_bytes)} bytes")

            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename=GoogleAdsCopyFilled.xlsx')
            self.end_headers()
            self.wfile.write(file_bytes)
        except Exception as e:
            self.send_response(500)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error = {"error": str(e)}
            self.wfile.write(json.dumps(error).encode())

    def do_OPTIONS(self):
        debug("Handling OPTIONS preflight")
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()